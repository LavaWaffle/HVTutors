import openpyxl

def getSheet(path):
	wb = openpyxl.load_workbook(path)
	sheet = wb.active
	return sheet

def getData(sheet):
	returnDict = {}
	
	maxRow = sheet.max_row
	
	for irow in range(2, maxRow+1):
		returnDict[sheet.cell(row = irow, column = 1).value] = {
			"name":sheet.cell(row=irow, column=2).value,
			"description":sheet.cell(row=irow, column = 3).value,
			"type":sheet.cell(row=irow, column = 4).value,
			"1Hour":int(sheet.cell(row=irow, column = 5).value),
			"90Mins":int(sheet.cell(row=irow, column = 6).value),
			"2Hours":int(sheet.cell(row=irow, column = 7).value),
		}

	return returnDict

def getType(bigDict, type):
	returnDict = {}
	for key, value in bigDict.items():
		for mkey, mvalue in value.items():
			if mvalue == type:
				returnDict[key] = value
		
	return returnDict

def getFront(bigDict, frontList):
	returnList = []
	for list in frontList:
		for key, value in bigDict.items():
			if key == list[0]:
				value['gId'] = list[-1]

				if list[1] == '1 hour':
					value['chosenTime'] = ['1 Hour', value['1Hour']]
				elif list[1] == '90 minutes':
					value['chosenTime'] = ['90 Minutes', value['90Mins']]
				else:
					value['chosenTime'] = ['2 Hours', value['2Hours']]
				returnList.append([key, value])
	return returnList